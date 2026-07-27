"""Geração da planilha Excel consolidada."""

from __future__ import annotations

import logging
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models import DependencyInfo

logger = logging.getLogger("dependency_report.report")

# Limiar definido no enunciado: dependências abaixo dele são destacadas.
SCORE_THRESHOLD = 65

DATE_FORMAT = "%Y-%m-%d"
HEADER_ROW = 1
FIRST_DATA_ROW = 2

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_HEIGHT = 34

ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ALERT_FONT = Font(color="9C0006")

TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


@dataclass(frozen=True)
class Column:
    """Descrição de uma coluna da planilha.

    Manter cabeçalho, extração do valor e largura juntos evita que a ordem das
    colunas e a ordem dos valores saiam de sincronia.
    """

    header: str
    extract: Callable[[DependencyInfo], Any]
    width: int
    numeric: bool = False


def _format_date(moment: datetime | None) -> str | None:
    """Formata a data como texto.

    O Excel não aceita datetime com fuso horário, e o formato ISO curto
    continua ordenável alfabeticamente na planilha.
    """
    return moment.strftime(DATE_FORMAT) if moment else None


COLUMNS: Sequence[Column] = (
    Column("Nome", lambda item: item.name, width=20),
    Column("Versão requisitada", lambda item: item.requested_version, width=18),
    Column("Última versão PyPI", lambda item: item.pypi_version, width=18),
    Column("Descrição", lambda item: item.description, width=50),
    Column("Licença", lambda item: item.license, width=20),
    Column("Última publicação", lambda item: _format_date(item.last_release_date), width=17),
    Column("Score Snyk", lambda item: item.snyk_score, width=11, numeric=True),
    Column(
        "Vulnerabilidades (total)",
        lambda item: item.vulnerabilities_total,
        width=15,
        numeric=True,
    ),
    Column(
        "Vulnerabilidades (versão atual)",
        lambda item: item.vulnerabilities_latest,
        width=15,
        numeric=True,
    ),
    Column(
        "Vulnerabilidades (PyPI/OSV)",
        lambda item: item.vulnerabilities_pypi,
        width=15,
        numeric=True,
    ),
    Column("Notas", lambda item: item.notes, width=32),
)


def build_report(dependencies: list[DependencyInfo], destination: Path) -> None:
    """Grava a planilha com as dependências e seus dados coletados."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dependências"

    _write_header(sheet)
    for offset, dependency in enumerate(dependencies):
        _write_row(sheet, FIRST_DATA_ROW + offset, dependency)

    _apply_layout(sheet, len(dependencies))
    _write_legend(workbook)
    workbook.save(destination)


def _write_header(sheet: Worksheet) -> None:
    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=column.header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    sheet.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT


def _write_row(sheet: Worksheet, row_index: int, dependency: DependencyInfo) -> None:
    below_threshold = _is_below_threshold(dependency.snyk_score)

    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=row_index, column=index, value=column.extract(dependency))

        if isinstance(cell.value, str):
            # Sem isto, valores como "==3.1.3" viram fórmula e a planilha
            # exibe #ERROR! ao abrir.
            cell.data_type = "s"

        cell.alignment = NUMBER_ALIGNMENT if column.numeric else TEXT_ALIGNMENT

        if below_threshold:
            cell.fill = ALERT_FILL
            cell.font = ALERT_FONT


def _is_below_threshold(score: float | None) -> bool:
    """Indica se o score exige destaque.

    Um pacote sem score não é destacado: a informação está ausente, o que não
    é o mesmo que uma nota ruim.
    """
    if score is None:
        return False

    try:
        return float(score) < SCORE_THRESHOLD
    except (TypeError, ValueError):
        logger.warning("Score em formato inesperado, linha não destacada: %r", score)
        return False


def _apply_layout(sheet: Worksheet, row_count: int) -> None:
    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = column.width

    last_column = get_column_letter(len(COLUMNS))
    # Congela o cabeçalho e habilita ordenação/filtro pelas colunas.
    sheet.freeze_panes = sheet.cell(row=FIRST_DATA_ROW, column=1)
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{last_column}{HEADER_ROW + row_count}"


def _write_legend(workbook: Workbook) -> None:
    """Cria uma aba explicando as colunas e o critério de destaque."""
    sheet = workbook.create_sheet("Legenda")
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 78

    descriptions = [
        ("Nome", "Nome da dependência conforme declarado no arquivo de entrada."),
        (
            "Versão requisitada",
            "Especificador declarado no projeto, como '>=3.0' ou '==2.6.1'.",
        ),
        ("Última versão PyPI", "Versão mais recente publicada no PyPI."),
        ("Descrição", "Resumo do pacote informado pelo PyPI."),
        ("Licença", "Licença declarada no PyPI."),
        ("Última publicação", "Data da publicação mais recente no PyPI."),
        ("Score Snyk", "Package Health Score do portal Snyk, de 0 a 100."),
        ("Vulnerabilidades (total)", "Todas as vulnerabilidades já registradas para o pacote."),
        (
            "Vulnerabilidades (versão atual)",
            "Quantas dessas ainda afetam a versão mais recente. É um subconjunto do total.",
        ),
        (
            "Vulnerabilidades (PyPI/OSV)",
            "Vulnerabilidades na versão atual segundo a base OSV, informada pela API do PyPI. "
            "É uma fonte independente do Snyk: divergência entre as duas colunas significa que "
            "as bases avaliam de forma diferente quais versões são afetadas.",
        ),
        ("Notas", "Motivo de eventual falha na coleta dos dados da dependência."),
    ]

    header = sheet.cell(row=1, column=1, value="Coluna")
    header.fill = HEADER_FILL
    header.font = HEADER_FONT
    meaning = sheet.cell(row=1, column=2, value="Significado")
    meaning.fill = HEADER_FILL
    meaning.font = HEADER_FONT

    for offset, (name, description) in enumerate(descriptions, start=2):
        sheet.cell(row=offset, column=1, value=name).font = Font(bold=True)
        sheet.cell(row=offset, column=2, value=description)

    legend_row = len(descriptions) + 3
    sheet.cell(row=legend_row, column=1, value="Destaque em vermelho").font = Font(bold=True)
    explanation = sheet.cell(
        row=legend_row,
        column=2,
        value=(
            f"Dependências com Score Snyk inferior a {SCORE_THRESHOLD}. "
            "Pacotes sem score não são destacados: o dado está ausente, "
            "o que não significa nota ruim."
        ),
    )
    for cell in (sheet.cell(row=legend_row, column=1), explanation):
        cell.fill = ALERT_FILL
        cell.font = Font(bold=cell.column == 1, color="9C0006")
