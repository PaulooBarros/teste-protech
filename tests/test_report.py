"""Testes da geração da planilha Excel."""

from datetime import datetime

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.models import DependencyInfo
from src.report import COLUMNS, build_report

DESTAQUE = "00FFC7CE"
SEM_PREENCHIMENTO = "00000000"
COLUNA_SCORE = 7


def gerar(tmp_path, dependencias):
    destino = tmp_path / "report.xlsx"
    build_report(dependencias, destino)
    return load_workbook(destino).active


class TestEstrutura:
    def test_gera_o_arquivo(self, tmp_path):
        destino = tmp_path / "report.xlsx"
        build_report([DependencyInfo(name="flask")], destino)

        assert destino.exists()

    def test_primeira_linha_e_o_cabecalho(self, tmp_path):
        sheet = gerar(tmp_path, [DependencyInfo(name="flask")])
        cabecalho = [cell.value for cell in sheet[1]]

        assert cabecalho[0] == "Nome"
        assert "Score Snyk" in cabecalho
        assert "Vulnerabilidades (total)" in cabecalho
        assert "Vulnerabilidades (versão atual)" in cabecalho

    def test_cabecalho_em_negrito(self, tmp_path):
        sheet = gerar(tmp_path, [DependencyInfo(name="flask")])

        assert all(cell.font.bold for cell in sheet[1])

    def test_uma_linha_por_dependencia(self, tmp_path):
        dependencias = [DependencyInfo(name="flask"), DependencyInfo(name="requests")]
        sheet = gerar(tmp_path, dependencias)

        assert sheet.max_row == len(dependencias) + 1  # + cabeçalho

    def test_grava_os_dados_coletados(self, tmp_path):
        dependencia = DependencyInfo(
            name="flask",
            requested_version="==3.1.3",
            pypi_version="3.1.3",
            description="Um framework web",
            license="BSD-3-Clause",
            last_release_date=datetime(2026, 2, 19, 5, 0, 57),
            snyk_score=90,
            vulnerabilities_total=6,
            vulnerabilities_latest=0,
            vulnerabilities_pypi=0,
        )
        sheet = gerar(tmp_path, [dependencia])
        linha = [cell.value for cell in sheet[2]]

        assert linha[:10] == [
            "flask",
            "==3.1.3",
            "3.1.3",
            "Um framework web",
            "BSD-3-Clause",
            "2026-02-19",
            90,
            6,
            0,
            0,
        ]

    def test_fontes_divergentes_aparecem_lado_a_lado(self, tmp_path):
        """A divergência entre Snyk e OSV é informação, não erro a esconder."""
        dependencia = DependencyInfo(
            name="pycrypto",
            vulnerabilities_total=4,
            vulnerabilities_latest=2,
            vulnerabilities_pypi=4,
        )
        sheet = gerar(tmp_path, [dependencia])

        assert [cell.value for cell in sheet[2]][7:10] == [4, 2, 4]

    def test_campos_ausentes_ficam_vazios(self, tmp_path):
        sheet = gerar(tmp_path, [DependencyInfo(name="desconhecido")])
        linha = [cell.value for cell in sheet[2]]

        assert linha[0] == "desconhecido"
        assert all(valor in (None, "") for valor in linha[1:])


class TestValoresDeTexto:
    """O Excel interpreta qualquer célula iniciada por `=` como fórmula."""

    @pytest.mark.parametrize("especificador", ["==3.1.3", "==2.6.1", "=1.0"])
    def test_especificador_com_igual_nao_vira_formula(self, tmp_path, especificador):
        dependencia = DependencyInfo(name="flask", requested_version=especificador)
        sheet = gerar(tmp_path, [dependencia])
        cell = sheet.cell(row=2, column=2)

        assert cell.data_type == "s"
        assert cell.value == especificador

    @pytest.mark.parametrize("especificador", [">=2.31.0", "<2.0", "~=5.0", "^4.17"])
    def test_demais_especificadores_seguem_como_texto(self, tmp_path, especificador):
        dependencia = DependencyInfo(name="requests", requested_version=especificador)
        sheet = gerar(tmp_path, [dependencia])

        assert sheet.cell(row=2, column=2).value == especificador


class TestUsabilidade:
    def test_cabecalho_fica_congelado(self, tmp_path):
        sheet = gerar(tmp_path, [DependencyInfo(name="flask")])

        assert sheet.freeze_panes == "A2"

    def test_habilita_filtro_nas_colunas(self, tmp_path):
        """O filtro deve cobrir todas as colunas e todas as linhas de dados."""
        dependencias = [DependencyInfo(name="flask"), DependencyInfo(name="requests")]
        sheet = gerar(tmp_path, dependencias)

        ultima_coluna = get_column_letter(len(COLUMNS))
        ultima_linha = len(dependencias) + 1  # + cabeçalho

        assert sheet.auto_filter.ref == f"A1:{ultima_coluna}{ultima_linha}"

    def test_planilha_tem_aba_de_legenda(self, tmp_path):
        destino = tmp_path / "report.xlsx"
        build_report([DependencyInfo(name="flask")], destino)
        workbook = load_workbook(destino)

        assert workbook.sheetnames == ["Dependências", "Legenda"]

    def test_legenda_descreve_todas_as_colunas(self, tmp_path):
        destino = tmp_path / "report.xlsx"
        build_report([DependencyInfo(name="flask")], destino)
        workbook = load_workbook(destino)

        dados = workbook["Dependências"]
        legenda = workbook["Legenda"]
        cabecalhos = {cell.value for cell in dados[1]}
        descritos = {cell.value for cell in legenda["A"]}

        assert cabecalhos <= descritos


class TestDestaqueVisual:
    """Requisito: destacar dependências com score inferior a 65."""

    @pytest.mark.parametrize("score", [0, 44, 64])
    def test_score_abaixo_de_65_e_destacado(self, tmp_path, score):
        sheet = gerar(tmp_path, [DependencyInfo(name="pycrypto", snyk_score=score)])

        assert sheet.cell(row=2, column=1).fill.start_color.rgb == DESTAQUE

    @pytest.mark.parametrize("score", [65, 90, 100])
    def test_score_a_partir_de_65_nao_e_destacado(self, tmp_path, score):
        sheet = gerar(tmp_path, [DependencyInfo(name="flask", snyk_score=score)])

        assert sheet.cell(row=2, column=1).fill.start_color.rgb == SEM_PREENCHIMENTO

    def test_score_ausente_nao_e_destacado(self, tmp_path):
        """Sem score não há como julgar o pacote: não deve parecer inseguro."""
        sheet = gerar(tmp_path, [DependencyInfo(name="desconhecido", snyk_score=None)])

        assert sheet.cell(row=2, column=1).fill.start_color.rgb == SEM_PREENCHIMENTO

    def test_a_linha_inteira_e_destacada(self, tmp_path):
        sheet = gerar(tmp_path, [DependencyInfo(name="pycrypto", snyk_score=44)])

        assert all(cell.fill.start_color.rgb == DESTAQUE for cell in sheet[2])

    def test_destaca_apenas_as_linhas_afetadas(self, tmp_path):
        dependencias = [
            DependencyInfo(name="flask", snyk_score=90),
            DependencyInfo(name="pycrypto", snyk_score=44),
            DependencyInfo(name="requests", snyk_score=80),
        ]
        sheet = gerar(tmp_path, dependencias)
        preenchimentos = [sheet.cell(row=i, column=1).fill.start_color.rgb for i in range(2, 5)]

        assert preenchimentos == [SEM_PREENCHIMENTO, DESTAQUE, SEM_PREENCHIMENTO]
