"""Testes do resumo estatístico da execução."""

import pytest
from openpyxl import load_workbook

from src.models import DependencyInfo
from src.report import ReportSummary, build_report

DESTAQUE = "00FFC7CE"
SEM_PREENCHIMENTO = "00000000"


def resumir(**campos):
    """Cria uma dependência com os campos informados e resume a lista."""
    return ReportSummary.from_dependencies([DependencyInfo(name="pacote", **campos)])


class TestContagens:
    def test_conta_as_dependencias_analisadas(self):
        dependencias = [DependencyInfo(name=nome) for nome in ("flask", "requests", "django")]

        assert ReportSummary.from_dependencies(dependencias).analyzed == 3

    def test_lista_vazia_zera_todos_os_numeros(self):
        resumo = ReportSummary.from_dependencies([])

        assert resumo.analyzed == 0
        assert resumo.below_threshold == 0
        assert resumo.current_vulnerabilities == 0

    @pytest.mark.parametrize("score, esperado", [(44, 1), (64, 1), (65, 0), (90, 0), (None, 0)])
    def test_conta_os_que_ficam_abaixo_do_limiar(self, score, esperado):
        assert resumir(snyk_score=score).below_threshold == esperado

    def test_conta_os_sem_score(self):
        """Pacote fora do catálogo do portal não é falha, mas precisa aparecer."""
        dependencias = [
            DependencyInfo(name="flask", snyk_score=90),
            DependencyInfo(name="obscuro", snyk_score=None),
        ]

        assert ReportSummary.from_dependencies(dependencias).without_score == 1

    def test_conta_falhas_de_coleta(self):
        dependencias = [
            DependencyInfo(name="flask"),
            DependencyInfo(name="quebrado", notes="tempo esgotado"),
        ]

        assert ReportSummary.from_dependencies(dependencias).collection_failures == 1


class TestPacotesVersusVulnerabilidades:
    """Quantos pacotes olhar é diferente de quantos problemas existem."""

    def test_um_pacote_muito_vulneravel_conta_como_um_pacote(self):
        resumo = resumir(vulnerabilities_latest=30)

        assert resumo.packages_at_risk == 1
        assert resumo.current_vulnerabilities == 30

    def test_soma_as_vulnerabilidades_de_todos(self):
        dependencias = [
            DependencyInfo(name="a", vulnerabilities_latest=2, vulnerabilities_total=10),
            DependencyInfo(name="b", vulnerabilities_latest=3, vulnerabilities_total=5),
        ]
        resumo = ReportSummary.from_dependencies(dependencias)

        assert resumo.packages_at_risk == 2
        assert resumo.current_vulnerabilities == 5
        assert resumo.historical_vulnerabilities == 15

    def test_zero_vulnerabilidades_nao_conta_como_risco(self):
        assert resumir(vulnerabilities_latest=0).packages_at_risk == 0

    def test_dado_ausente_nao_conta_como_risco(self):
        """Não saber é diferente de saber que está limpo."""
        assert resumir(vulnerabilities_latest=None).packages_at_risk == 0


class TestDivergenciaEntreFontes:
    def test_conta_quando_snyk_e_osv_discordam(self):
        resumo = resumir(vulnerabilities_latest=2, vulnerabilities_pypi=4)

        assert resumo.source_disagreements == 1

    def test_nao_conta_quando_concordam(self):
        resumo = resumir(vulnerabilities_latest=0, vulnerabilities_pypi=0)

        assert resumo.source_disagreements == 0

    @pytest.mark.parametrize(
        "snyk, osv",
        [(None, 4), (2, None), (None, None)],
    )
    def test_dado_ausente_nao_e_discordancia(self, snyk, osv):
        resumo = resumir(vulnerabilities_latest=snyk, vulnerabilities_pypi=osv)

        assert resumo.source_disagreements == 0


class TestAbaNaPlanilha:
    def gerar(self, tmp_path, dependencias):
        destino = tmp_path / "report.xlsx"
        build_report(dependencias, destino)
        return load_workbook(destino)

    def test_ordem_das_abas(self, tmp_path):
        """A lista exigida vem primeiro: é o que se deve encontrar ao abrir."""
        workbook = self.gerar(tmp_path, [DependencyInfo(name="flask")])

        assert workbook.sheetnames == ["Dependências", "Resumo", "Legenda"]

    def test_grava_os_numeros(self, tmp_path):
        dependencias = [
            DependencyInfo(name="flask", snyk_score=90, vulnerabilities_latest=0),
            DependencyInfo(name="pycrypto", snyk_score=44, vulnerabilities_latest=2),
        ]
        sheet = self.gerar(tmp_path, dependencias)["Resumo"]
        valores = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}

        assert valores["Dependências analisadas"] == 2
        assert valores["Com score abaixo de 65"] == 1
        assert valores["Pacotes com vulnerabilidade na versão atual"] == 1

    def test_destaca_a_linha_quando_ha_problema(self, tmp_path):
        dependencias = [DependencyInfo(name="pycrypto", snyk_score=44)]
        sheet = self.gerar(tmp_path, dependencias)["Resumo"]
        linha = next(
            row for row in sheet.iter_rows(min_row=2) if row[0].value == "Com score abaixo de 65"
        )

        assert linha[0].fill.start_color.rgb == DESTAQUE

    def test_nao_destaca_quando_o_valor_e_zero(self, tmp_path):
        """Zero problemas é boa notícia: não deve parecer alerta."""
        dependencias = [DependencyInfo(name="flask", snyk_score=90)]
        sheet = self.gerar(tmp_path, dependencias)["Resumo"]
        linha = next(
            row for row in sheet.iter_rows(min_row=2) if row[0].value == "Com score abaixo de 65"
        )

        assert linha[0].fill.start_color.rgb == SEM_PREENCHIMENTO
